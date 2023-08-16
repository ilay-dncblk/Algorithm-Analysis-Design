import openpyxl
import time
import matplotlib.pyplot as plt
import numpy as np

def knapsack(filename):
    with open(filename, 'r') as f:
        lines = f.readlines()

    num_items, total_weight = map(int, lines[0].split())

    values = []
    weights = []

    for line in lines[1:]:
        v, w = map(int, line.split())
        values.append(v)
        weights.append(w)

    table = [[0 for _ in range(total_weight+1)] for _ in range(num_items+1)]

    for i in range(1, num_items+1):
        for j in range(1, total_weight+1):
            if weights[i-1] > j:
                table[i][j] = table[i-1][j]
            else:
                table[i][j] = max(table[i-1][j], table[i-1][j-weights[i-1]] + values[i-1])

    max_value = table[-1][-1]
    
    selected_items = [0] * num_items
    selected_item_values = [0] * num_items
    j = total_weight
    for i in range(num_items, 0, -1):
        if table[i][j] != table[i-1][j]:
            selected_item_values[i-1] = values[i-1]
            selected_items[i-1] = 1
            j -= weights[i-1]

    selected_item_values = [x for x in selected_item_values if x != 0]
    return max_value, selected_items, selected_item_values


file_names = ['ks_4_0.txt', 'ks_19_0.txt', 'ks_200_0.txt','ks_10000_0.txt']
file_item_counts = [4, 19, 200, 10000]

wb = openpyxl.Workbook()
ws = wb.active

ws.cell(row=1, column=1, value='Öğrenci Numarası:212802040, Ad Soyad: İlayda DİNÇBİLEK')

ws.cell(row=2, column=1, value='Dosya Boyut')
ws.cell(row=2, column=2, value='Optimal Value Değeri')
ws.cell(row=2, column=3, value='Optimal çözüm(itemler arasında SADECE bir boşluk bırakılmalı, ‘.’, ‘,’, ’-’ vb. karakterler kullanılmamalıdır)')
ws.cell(row=2, column=4, value='Optimal çözüme dahil edilen itemler')

for i, file_name in enumerate(file_names):
    start_time = time.time()
    max_value, selected_items,selected_item_values = knapsack(file_name)
    ws.cell(row=i+3, column=1, value=file_item_counts[i])
    ws.cell(row=i+3, column=2, value=max_value)
    ws.cell(row=i+3, column=3, value= ' '.join(str(x) for x in selected_item_values))
    ws.cell(row=i+3, column=4, value=' '.join(str(x) for x in selected_items))
    
    end_time = time.time()
    exec_time = end_time - start_time
    print('Total time: {}'.format(exec_time), 'seconds',"File name: ",file_name)
    x = np.array(file_item_counts[i])
    y = np.array(exec_time)
    
    
    plt.plot(x,y,'o')
    
    
    

wb.save('knapsack_results.xlsx')

plt.xlabel("File Item Count")
plt.ylabel("Exec Time")
plt.title("Total Value and Exec Time Graph")
plt.show()
plt.savefig("Total Value and Exec Time Graph.png")

